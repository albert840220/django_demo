from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.utils.decorators import method_decorator
from django.views import View

from electrodes.forms import ProfileForm, form_validation_error, CalibrationForm, InspectionForm, PmScheduleForm, \
    CustomerForm, FactoryForm, EqptPhForm, EqptCodForm, CustomerSurveyForm, BusinessScheduleForm
from electrodes.models import Profile, Calibration, Inspection, Factory, EqptPh, EqptCod, Customer, PmSchedule, \
    Staff, CustomerSurvey, BusinessSchedule

# 多條件查詢圖表
from plotly.offline import plot
import plotly.graph_objs as go
from django.db import connection
import pandas as pd

# 下載excel
import openpyxl
from django.utils.http import urlquote
from django.http import HttpResponse

from django.conf import settings
# decorators
from electrodes.decorators import role_required
import json
from django.core import serializers
import datetime as dt
import dateutil.relativedelta
from django import forms

def form_business_schedule(request):
    if request.method == 'POST':
        form = BusinessScheduleForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("/business")
    else:
        form = BusinessScheduleForm()
    return render(request, 'add_business.html',{"form": form})

def tables_business_schedule(request):
    business = BusinessSchedule.objects.all()
    business_list = [{
        "id": r.id,
        "staff_name": r.staff_id.name,
        "visit_date": str(r.visit_date),
        "factory_id": r.factory_id.name,
        "customer_id": r.customer_id.customer_name,
        "purpose_of_visit": r.purpose_of_visit,
    } for r in business]
    return render(request, "tables_business_schedule.html", {"business_schedule": json.dumps(business_list)})


def tables_customer_survey(request):
    survey = CustomerSurvey.objects.all()
    survey_list = [{
        "id": r.id,
        "company": r.form_id.factory_id.name,
        "purpose_of_visit": r.form_id.purpose_of_visit,
    } for r in survey]
    return render(request, "tables_customer_survey.html", {"survey": json.dumps(survey_list)})


def display_png(request, id):
    survey = CustomerSurvey.objects.get(id=id)
    business = BusinessSchedule.objects.get(id=survey.form_id.id)
    return render(request, 'display_sign.html',{"survey":survey,"business": business})


def form_customer_survey(request):
    # form = SignatureForm(request.POST or None)
    if request.method == 'POST':
        form = CustomerSurveyForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("/survey")
            # signature = form.cleaned_data.get('signature')
            # if signature:
            #     # as an image
            #     signature_picture = draw_signature(signature)
            #     print(signature_picture)
            #     # or as a file
            #     signature_file_path = draw_signature(signature, as_file=True)
            #     print(signature_file_path)
    else:
        form = CustomerSurveyForm()
    return render(request, 'add_signature.html',{"form": form})


def default_work_schedule(request):
    if request.method == 'POST':
        form = PmScheduleForm(request.POST)
        if form.is_valid():
            return render(request, 'add_job.html', {"form": form})
    else:
        form = PmScheduleForm()
    return render(request, 'add_job.html', {"form": form})


def close_case(request, id):
    """
    結案功能
    """
    schedule = PmSchedule.objects.get(id=id)
    sn = schedule.equipment_sn
    if request.method == 'POST':
        if EqptPh.objects.filter(sn=sn).exists():
            eqpt_data = EqptPh.objects.get(sn=sn)
            form = EqptPhForm(request.POST, instance=eqpt_data)
        elif EqptCod.objects.filter(sn=sn).exists():
            eqpt_data = EqptCod.objects.get(sn=sn)
            form = EqptCodForm(request.POST, instance=eqpt_data)
            print(form)
        if form.is_valid():
            form.save()
            print('alre save')
            if EqptPh.objects.filter(sn=sn).exists():
                EqptPh.objects.filter(sn=sn).update(state='0')
            elif EqptCod.objects.filter(sn=sn).exists():
                EqptCod.objects.filter(sn=sn).update(state='0')
            PmSchedule.objects.filter(equipment_sn=sn).update(state='1')
            return redirect("/schedule")
    else:

        if EqptPh.objects.filter(sn=sn).exists():
            eqpt_data = EqptPh.objects.get(sn=schedule.equipment_sn)
            form = EqptPhForm()
            form = EqptPhForm(
                initial={'factory_id': eqpt_data.factory_id,"model": eqpt_data.model,"sn": eqpt_data.sn, "standard_a": eqpt_data.standard_a,
                         "standard_b": eqpt_data.standard_b, "sv1_ph": eqpt_data.sv1_ph,
                         "sv2_ph": eqpt_data.sv2_ph, "pump_ph": eqpt_data.pump_ph})# ,"state": eqpt_data.state})
            form.fields['factory_id'].widget = forms.HiddenInput()
            form.fields['model'].widget = forms.HiddenInput()
            form.fields['sn'].widget = forms.HiddenInput()
            # form.fields['state'].widget = forms.HiddenInput()
            form.fields['last_pm'].label = "實際保養日"

        elif EqptCod.objects.filter(sn=sn).exists():
            eqpt_data = EqptCod.objects.get(sn=schedule.equipment_sn)
            form = EqptCodForm()
            form = EqptCodForm(
                initial={'factory_id': eqpt_data.factory_id,"model": eqpt_data.model,"sn": eqpt_data.sn,"standard_c": eqpt_data.standard_c,
                         "standard_d": eqpt_data.standard_d, "sv1_cod": eqpt_data.sv1_cod,
                         "sv2_cod": eqpt_data.sv2_cod, 'sv3_cod': eqpt_data.sv3_cod,
                         "pump_cod": eqpt_data.pump_cod,"state": eqpt_data.state})
            form.fields['factory_id'].widget = forms.HiddenInput()
            form.fields['model'].widget = forms.HiddenInput()
            form.fields['sn'].widget = forms.HiddenInput()
            form.fields['state'].widget = forms.HiddenInput()
            form.fields['last_pm'].label = "實際保養日"

    return render(request, 'close_case.html', {"form": form,"schedule":schedule,"sn":sn})


def add_equipment_ph(request):
    """
    新增pH機台
    """
    if request.method == 'POST':
        form = EqptPhForm(request.POST)
        if form.is_valid():
            factory_id = form.cleaned_data['factory_id']
            model = form.cleaned_data['model']
            sn = form.cleaned_data['sn']
            last_pm = form.cleaned_data['last_pm']
            standard_a = form.cleaned_data['standard_a']
            standard_b = form.cleaned_data['standard_b']
            sv1_ph = form.cleaned_data['sv1_ph']
            sv2_ph = form.cleaned_data['sv2_ph']
            pump_ph = form.cleaned_data['pump_ph']
            last_pm_staff = form.cleaned_data['last_pm_staff']
            # state = form.cleaned_data['state']
            EqptPh.objects.create(factory_id=factory_id,model=model,sn=sn,last_pm=last_pm,standard_a=standard_a,standard_b=standard_b,sv1_ph=sv1_ph,sv1_ph_lifetime=1,sv2_ph=sv2_ph,sv2_ph_lifetime=1,pump_ph=pump_ph,pump_ph_lifetime=1,last_pm_staff=last_pm_staff,state=0)
            return redirect("/equipment/ph")
    else:
        form = EqptPhForm()
    return render(request, 'add_eqpt_ph.html', {"form": form})


def add_customer(request):
    """
    新增客戶
    """
    if request.method == 'POST':
        form_factory = FactoryForm(request.POST)
        form_customer = CustomerForm(request.POST)
        if form_factory.is_valid() and form_customer.is_valid():
            factory_name = form_factory.cleaned_data['name']
            city = form_factory.cleaned_data['city']
            address = form_factory.cleaned_data['address']
            Factory.objects.create(name=factory_name, city=city, address=address)
            new_one = Factory.objects.last()
            customer_name = form_customer.cleaned_data['customer_name']
            email = form_customer.cleaned_data['email']
            phone = form_customer.cleaned_data['phone']
            Customer.objects.create(factory_id=new_one,customer_name=customer_name,email=email,phone=phone)
            return redirect('/customer')
    else:
        form_customer = CustomerForm()
        form_factory = FactoryForm()
    return render(request, 'add_customer.html', {"form_factory": form_factory,"form_customer":form_customer})


def tables_pm_schedule(request):
    """
    工作排程表
    """
    pm_schedule = PmSchedule.objects.filter(state=0)
    schedule_list = [{
        "id": r.id,
        "next_pm_time": str(r.next_maintenance_planned_on),
        "customer_name": r.customer_id.customer_name,
        "factory_name": r.factory_id.name,
        "model": r.equipment_model,
        "sn": r.equipment_sn,
        "description": r.description,
        "staff_name": r.staff_id.name,
    } for r in pm_schedule]
    return render(request, "tables_pm_schedule.html", {"schedule_list": json.dumps(schedule_list)})


def add_job(request, sn):
    """
    新增派工單
    """
    if sn == 'new':
        form = PmScheduleForm()
    else:
        if EqptPh.objects.filter(sn=sn).exists():
            eqpt_data = EqptPh.objects.get(sn=sn)
            form = PmScheduleForm(initial={"next_maintenance_planned_on": eqpt_data.last_pm, "factory_id": eqpt_data.factory_id,
                                       "equipment_model": eqpt_data.model, 'equipment_sn': eqpt_data.sn})
        elif EqptCod.objects.filter(sn=sn).exists():
            eqpt_data = EqptCod.objects.get(sn=sn)
            form = PmScheduleForm(
                initial={"next_maintenance_planned_on": eqpt_data.last_pm, "factory_id": eqpt_data.factory_id,
                         "equipment_model": eqpt_data.model, 'equipment_sn': eqpt_data.sn})
    form.fields["staff_id"].queryset =Staff.objects.filter(department_id=3) # 選項只有技術部的同仁
    print(f'SN:{sn}')
    if request.method == 'POST':
        form = PmScheduleForm(request.POST)
        if form.is_valid():
            next_maintenance_planned_on = form.cleaned_data['next_maintenance_planned_on']
            customer_id = form.cleaned_data['customer_id']
            factory_id = form.cleaned_data['factory_id']
            equipment_model = form.cleaned_data['equipment_model']
            equipment_sn = form.cleaned_data['equipment_sn']
            description = form.cleaned_data['description']
            staff_id = form.cleaned_data['staff_id']
            PmSchedule.objects.create(next_maintenance_planned_on=next_maintenance_planned_on, customer_id=customer_id,
                                      factory_id=factory_id, equipment_model=equipment_model, equipment_sn=equipment_sn,
                                      description=description, staff_id=staff_id, state=0)
            if EqptPh.objects.filter(sn=sn).exists():
                EqptPh.objects.filter(sn=sn).update(state='1')
            elif EqptCod.objects.filter(sn=sn).exists():
                EqptCod.objects.filter(sn=sn).update(state='1')
            elif sn == 'new':
                print('in new')
                if EqptPh.objects.filter(sn=equipment_sn).exists():
                    EqptPh.objects.filter(sn=equipment_sn).update(state='1')
                elif EqptCod.objects.filter(sn=equipment_sn).exists():
                    EqptCod.objects.filter(sn=equipment_sn).update(state='1')
            return redirect("/schedule")
    return render(request, 'add_job.html', {"sn": sn, "form": form})


def equipment_type(request):
    """
    機台
    """
    return render(request, "equipment_type.html")


def tables_pre_repair(request):
    """
    預知保養項目
    """
    today = dt.date.today()
    after_one_month = (dt.datetime.strptime(str(today), "%Y-%m-%d") + dateutil.relativedelta.relativedelta(
        days=30)).date()
    cod = EqptCod.objects.filter(last_pm__range=[today, after_one_month], state=0)
    print(f"Date range:{today} ~ {after_one_month}")
    eqpt_cod_list = [{
        "factory_id": r.factory_id.name,
        "model": r.model,
        "sn": r.sn,
        "last_pm": str(r.last_pm),
        # "standard_c": r.days_used_for_standard_c(),
        # "standard_d": r.days_used_for_standard_d(),
        # "sv1_cod": r.days_used_for_sv1_cod(),
        # "sv2_cod": r.days_used_for_sv2_cod(),
        # "sv3_cod": r.days_used_for_sv2_cod(),
        # "pump_cod": r.days_used_for_pump_cod(),
        # "sv1_ph_lifetime": r.sv1_lifetime(),
        # "sv2_ph_lifetime": r.sv2_lifetime(),
        # "sv3_ph_lifetime": r.sv2_lifetime(),
        # "state": r.state,
    } for r in cod]
    ph = EqptPh.objects.filter(last_pm__range=[today, after_one_month],state=0)
    eqpt_ph_list = [{
        "factory_id": r.factory_id.name,
        "model": r.model,
        "sn": r.sn,
        "last_pm": str(r.last_pm),
        # "standard_a": r.days_used_for_standard_a(),
        # "standard_b": r.days_used_for_standard_b(),
        # "sv1_ph": r.days_used_for_sv1_ph(),
        # "sv2_ph": r.days_used_for_sv2_ph(),
        # "pump_ph": r.days_used_for_pump_ph(),
        # "sv1_ph_lifetime": r.sv1_lifetime(),
        # "sv2_ph_lifetime": r.sv2_lifetime(),
        # "state": r.state,
    } for r in ph]
    all_eqpt = eqpt_cod_list + eqpt_ph_list
    return render(request, "tables_pre_repair.html", {"all_eqpt": json.dumps(all_eqpt)})


def tables_customer(request):
    """
    客戶資料
    """
    custom = Customer.objects.all()
    customer_list = [{
        "factory_name": r.factory_id.name,
        "address": r.show_address(),
        "name": r.customer_name,
        "phone": r.phone,
        "email": r.email,
    } for r in custom]
    return render(request, "tables_customer.html", {"customer_list": json.dumps(customer_list)})


def equipment_ph_detail(request, sn):
    """
    機台詳細資料
    """
    today = dt.date.today()
    eqpt_ph = EqptPh.objects.get(sn=sn)
    return render(request, 'equipment_ph_detail.html', {"eqpt_ph": eqpt_ph,'today':today})


def days_used(request):
    """
    零件時數表，顯示數值為"已使用天數"
    """
    # column_names = ['預計保養日期',"客戶名稱","廠商名稱","機型","機台序號","問題情況","作業者"]
    column_names = ["廠商名稱", "機型", "機台序號", '上次保養日', "標準液A使用天數", "標準液B使用天數", "SV1使用天數", "SV2使用天數", "Pump使用天數"]

    records = EqptPh.objects.all()
    eqpt_ph_list = [{
        "factory_id": r.factory_id.name,
        "model": r.model,
        "sn": r.sn,
        "last_pm": str(r.last_pm),
        "standard_a": r.days_used_for_standard_a(),
        "standard_b": r.days_used_for_standard_b(),
        "sv1_ph": r.days_used_for_sv1_ph(),
        "sv2_ph": r.days_used_for_sv2_ph(),
        "pump_ph": r.days_used_for_pump_ph(),
        "sv1_ph_lifetime": r.sv1_lifetime(),
        "sv2_ph_lifetime": r.sv2_lifetime(),
        "state": r.state,
    } for r in records]
    return render(request, 'calibrations-day.html', {"eqpt_ph_list": json.dumps(eqpt_ph_list)})


def tables_ph_parts(request):
    """
    零件時數表，顯示數值為"預計更換日"
    """
    # column_names = ['預計保養日期',"客戶名稱","廠商名稱","機型","機台序號","問題情況","作業者"]
    # column_names = ["廠商名稱", "機型", "機台序號", '上次保養日', "標準液A使用天數", "標準液B使用天數", "SV1使用天數", "SV2使用天數", "Pump使用天數"]
    records = EqptPh.objects.all()
    today = dt.datetime.today()
    after_14_days = (today + dateutil.relativedelta.relativedelta(days=14)).date()
    after_28_days = (today + dateutil.relativedelta.relativedelta(days=28)).date()
    print(today)
    print(after_14_days)
    print(after_28_days)
    eqpt_ph_list = [{
        "factory_id": r.factory_id.name,
        "model": r.model,
        "sn": r.sn,
        "last_pm": str(r.last_pm),
        "standard_a": r.days_used_for_standard_a(),
        "standard_b": r.days_used_for_standard_b(),
        "sv1_ph": r.days_used_for_sv1_ph(),
        "sv2_ph": r.days_used_for_sv2_ph(),
        "pump_ph": r.days_used_for_pump_ph(),
        "sv1_ph_lifetime": r.sv1_lifetime(),
        "sv2_ph_lifetime": r.sv2_lifetime(),
        "state": r.state,
    } for r in records]
    return render(request, 'tables_ph_parts.html',
                  {"eqpt_ph_list": json.dumps(eqpt_ph_list), "today": str(today.date()),
                   "after_14_days": str(after_14_days), "after_28_days": str(after_28_days)})


@login_required(login_url="/login/")
@role_required(allowed_roles=['技術部', '發展課', '網站管理員', '人事', '未定義'])
def welcome(request):
    if request.user.groups.all()[0].name == '未定義':
        return redirect('/call_it')
    return render(request, 'welcome.html')  # 必須用 render ，如使用redirect會一直重複指向自己


@login_required(login_url="/login/")
@role_required(allowed_roles=['技術部', '發展課', '網站管理員'])
def tables_calibration(request):
    """
    校正紀錄查詢
    """
    advanced_groups = ['技術部', '網站管理員']
    groupIsOK = request.user.groups.all()[0].name in advanced_groups  # 原html判斷式
    # 欄位名稱
    basic_column = ['校正日期', '電極序號']
    advanced_column = ['RFID', 'Slope', 'Offset']
    column_names = basic_column + advanced_column * (groupIsOK)
    # 欄位內容
    basic_fields = ['c_datetime', 'sensor_sn']  # 大家都可以看的欄位
    advanced_fields = ['rfid', 'slope', 'offset']  # ok_groups 可以看的欄位
    final_fields = basic_fields + advanced_fields * (groupIsOK)
    records = Calibration.objects.values_list(*final_fields)
    return render(request, "tables_calibrations.html", {'records': records, "column_names": column_names})


@login_required(login_url="/login/")
@role_required(allowed_roles=['未定義'])
# TODO 建立使用者時，自訂預設群組
def call_it(request):
    return render(request, 'call_it.html')


@login_required(login_url="/login/")
@role_required(allowed_roles=['技術部', '發展課', '網站管理員'])
def chartjs(request):
    """
    折線圖, chart-js寫法
    """
    labels = []
    data = []
    data1 = Calibration.objects.order_by('c_datetime')[:20].values_list('c_datetime', 'Offset', 'Slope')
    labels, offset, slope = zip(*data1)  # zip出來為tuple
    labels = list(i.strftime("%Y-%m-%d %H:%M:%S") for i in labels)
    offset = [float(i) for i in offset]
    slope = [float(i) for i in slope]
    return render(request, 'chartjs.html', {'labels': labels, 'offset': offset, 'slope': slope})


@login_required(login_url="/login/")
@role_required(allowed_roles=['技術部', '發展課', '網站管理員'])
def multi_condition_plot(request):
    """多條件查詢 + 畫折線圖"""
    start_date = None
    end_date = None
    plot_div = None
    # items = ['Slope','Offset']
    calibrations = Calibration.objects.values('sensor_sn').distinct()  # autocomplete 的項目
    if request.method == 'POST':
        if request.POST['sn_number']:
            sn_number = request.POST['sn_number']
            cursor = connection.cursor()
            query = f"""SELECT * FROM calibrations where sensor_sn='{sn_number}'"""
            df = pd.read_sql_query(query, connection)
            df = df.sort_values(by=['c_datetime'])
            print(f"sn_number: {df}")
            if df.empty:
                remind = '找不到符合條件的資料'
                return render(request, "trends.html", {'remind': remind, 'calibrations': calibrations})
            # 開始時間
            if request.POST['start_time']:
                start_date = request.POST['start_time']
            else:
                start_date = '2000-01-01'
            # 結束時間
            if request.POST['end_time']:
                end_date = request.POST['end_time']
            else:
                end_date = '2080-01-01'
            df = df[(df['c_datetime'] > f'{start_date}') & (df['c_datetime'] < f'{end_date}')]
            # up down show
            data = go.Scatter(x=df['c_datetime'], y=df['slope'],
                              mode='lines', name='test',
                              opacity=0.8, marker_color='green')
            layout = go.Layout(height=400, title='Slope',
                               legend=dict(x=0.4, y=-0.3, traceorder='normal', font=dict(size=12, ))
                               )
            fig = go.Figure(data=data, layout=layout)
            # 單線 顯示legend
            fig['data'][0]['showlegend'] = True
            fig['data'][0]['name'] = f"{request.POST['sn_number']}"
            fig.update_layout(title_x=0.5, hovermode='x')
            plot_div = plot(fig, output_type='div', config={'scrollZoom': True})

            data = go.Scatter(x=df['c_datetime'], y=df['offset'],
                              mode='lines', name='test',
                              opacity=0.8, marker_color='red')
            layout = go.Layout(height=300, title='offset')
            fig = go.Figure(data=data, layout=layout)
            fig.update_layout(title_x=0.5, hovermode='x')
            plot_div1 = plot(fig, output_type='div', config={'scrollZoom': True})
            # 兩線合併
            Slope = go.Scatter(x=df['c_datetime'], y=df['slope'],
                               mode='lines', name='slope',
                               opacity=0.8, marker_color='green')
            Offset = go.Scatter(x=df['c_datetime'], y=df['offset'],
                                mode='lines', name='offset',
                                opacity=0.8, marker_color='red')
            data = [Slope, Offset]
            layout = go.Layout(height=400, title=f"{request.POST['sn_number']} : Slope and Offset")
            fig = go.Figure(data=data, layout=layout)
            fig.update_layout(title_x=0.5, hovermode='x')
            plot_compare = plot(fig, output_type='div', config={'scrollZoom': True})

            return render(request, "trends.html",
                          {'plot_div': plot_div, 'plot_div1': plot_div1, 'plot_compare': plot_compare,
                           'calibrations': calibrations})  # ,'items':items})
    return render(request, "trends.html", {'calibrations': calibrations})  # , {'items':items})


@login_required(login_url="/login/")
@role_required(allowed_roles=['技術部', '網站管理員'])
def report_write(request):
    """填寫檢測報告書"""
    if request.method == "POST":
        form = InspectionForm(request.POST)
        if form.is_valid():
            # form_id = form.cleaned_data['form_id']
            delivery_date = form.cleaned_data['delivery_date']
            client = form.cleaned_data['client']
            incident_date = form.cleaned_data['incident_date']
            lifetime = form.cleaned_data['lifetime']
            hswe_name = form.cleaned_data['hswe_name']
            online_date = form.cleaned_data['online_date']
            # model = form.cleaned_data['model']
            sn = form.cleaned_data['sn']
            return_date = form.cleaned_data['return_date']
            system = form.cleaned_data['system']
            # tag_no = form.cleaned_data['tag_no']
            start_date = form.cleaned_data['start_date']
            # end_date = form.cleaned_data['end_date']
            # error_report = form.cleaned_data['error_report']
            sample_value = form.cleaned_data['sample_value']
            sample_conductivity = form.cleaned_data['sample_conductivity']
            sample_pressure = form.cleaned_data['sample_pressure']
            sample_temperature = form.cleaned_data['sample_temperature']
            install_type = form.cleaned_data['install_type']
            avg_life = form.cleaned_data['avg_life']
            Inspection.objects.create(form_id=str(incident_date).replace('-', '') + '-' + sn,
                                      delivery_date=delivery_date, client=client,
                                      incident_date=incident_date, lifetime=lifetime, hswe_name=hswe_name,
                                      online_date=online_date, sn=sn,
                                      return_date=return_date, system=system, start_date=start_date,
                                      sample_value=sample_value, sample_conductivity=sample_conductivity,
                                      sample_pressure=sample_pressure, sample_temperature=sample_temperature,
                                      install_type=install_type, avg_life=avg_life)
            return render(request, 'form_success.html')
    else:
        form = InspectionForm()
    return render(request, 'report_write.html', {'form': form})


@login_required(login_url="/login/")
@role_required(allowed_roles=['技術部', '網站管理員'])
def exportToExcel(request, form_id):
    """匯出檢測報告書"""
    inspection = Inspection.objects.get(form_id=form_id)
    wb = openpyxl.load_workbook(filename=f'{settings.BASE_DIR}/檢測與檢討報告書.xlsx')
    sheet = wb.worksheets[0]
    sheet.cell(row=4, column=4, value=f"{inspection.form_id}")  # 報告書編號
    sheet.cell(row=4, column=10, value=f"{inspection.delivery_date}")  # 出貨日期
    sheet.cell(row=5, column=4, value=f"{inspection.client}")  # 業主/現場連絡人
    sheet.cell(row=5, column=10, value=f"{inspection.incident_date}")  # 發生日期
    sheet.cell(row=5, column=13, value=f"{inspection.lifetime}")  # 使用壽命
    sheet.cell(row=6, column=4, value=f"{inspection.hswe_name}")  # 負責人
    sheet.cell(row=6, column=10, value=f"{inspection.online_date}")  # 上線日期
    sheet.cell(row=7, column=4, value=f"{inspection.sn}")  # 型號&序號
    sheet.cell(row=7, column=10, value=f"{inspection.return_date}")  # 取回日期
    sheet.cell(row=8, column=4, value=f"{inspection.system}")  # 系統&TagNO
    sheet.cell(row=8, column=10, value=f"{inspection.start_date}")  # 檢驗日期
    sheet.cell(row=11, column=6, value=f"{inspection.sample_value}")  # ph值
    sheet.cell(row=11, column=11, value=f"{inspection.sample_conductivity}")  # 導電度
    sheet.cell(row=12, column=6, value=f"{inspection.sample_pressure}")  # 壓力
    sheet.cell(row=12, column=11, value=f"{inspection.sample_temperature}")  # 溫度
    sheet.cell(row=13, column=6, value=f"{inspection.install_type}")  # 安裝方式
    sheet.cell(row=13, column=11, value=f"{inspection.avg_life}")  # 電極平均壽命
    # 設定響應頭
    response = HttpResponse(content_type='application/msexcel')
    # 設定下載檔案編碼，需要使用urlquote
    filename = urlquote(f'{inspection.form_id}.xlsx')
    response[
        'Content-Disposition'] = f"attachment;filename*=utf-8'zh_cn'{filename}"
    # 儲存Excel到相應中
    wb.save(response)
    return response


@login_required(login_url="/login/")
@role_required(allowed_roles=['技術部', '網站管理員'])
def tables_report(request):
    """檢測報告書查詢"""
    records = Inspection.objects.values('form_id', 'hswe_name')
    return render(request, "tables_report.html", {'records': records})
